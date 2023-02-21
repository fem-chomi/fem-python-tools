# Copyright (C) 2023 fem. All rights reserved.
import os
import glob
import chardet
import openpyxl
import hashlib
from full_text_search_dao import Dao

class Logic:
    def __init__(self):
        self.dao = Dao()

    def detect_file(self, filename):
        with open(filename, mode='rb') as f:
            bin = f.read()
            char_set = chardet.detect(bin)
            if char_set == None:
                return None
            else:
                return char_set['encoding']
    
    def read_text_file_utf8(self, filename):
        enc = self.detect_file(filename)
        try:
            if enc == None:
                return None
            elif enc == 'Windows-1252':
                with open(filename, mode='r', encoding='Shift_JIS') as f:
                    return f.read()
            else:
                with open(filename, mode='r', encoding=enc) as f:
                    return f.read()
        except UnicodeDecodeError:
            print('UnicodeDecodeError: {0}'.format(filename))

    def read_lines_text_file_utf8(self, filename):
        enc = self.detect_file(filename)
        try:
            if enc == None:
                return None
            elif enc == 'Windows-1252':
                with open(filename, mode='r', encoding='Shift_JIS') as f:
                    lines = []
                    for line in f.readlines():
                        lines.append(line)
                    return lines
            else:
                with open(filename, mode='r', encoding=enc) as f:
                    return f.read()
        except UnicodeDecodeError:
            print('UnicodeDecodeError: {0}'.format(filename))

    def registCommon(self, source_type, text, check):
        hash = hashlib.sha1(text.encode('utf-8')).hexdigest()
        if check:
            if self.dao.isExistTextTable(source_type, hash) == False:
                self.dao.insertTextTable(source_type, hash, text)
                return hash
            else:
                return None
        else:
            self.dao.insertTextTable(source_type, hash, text)
            return hash
        
    def registMemo(self, text):
        self.registCommon('memo', text, True)
        
    def registWWW(self, text, url):
        hash = self.registCommon('www', text, True)
        if hash != None:
            self.dao.insertWwwTable(hash, url)
    
    def registSourceCode(self, filename):
        if not os.path.exists(filename):
            return
        line_no = 0
        lines = self.read_lines_text_file_utf8(filename)
        if lines == None:
            return
        for line in lines:
            line_no = line_no + 1
            if line.strip() == '':
                continue
            hash = self.registCommon('code', line.strip(), False)
            if hash != None:
                self.dao.insertSourceFileTable(hash, filename, line_no)
    
    def registTextFile(self, filename):
        if not os.path.exists(filename):
            return
        line_no = 0
        lines = self.read_lines_text_file_utf8(filename)
        if lines == None:
            return
        for line in lines:
            line_no = line_no + 1
            if line.strip() == '':
                continue
            hash = self.registCommon('text', line.strip(), False)
            if hash != None:
                self.dao.insertTextFileTable(hash, filename, line_no)
    
    def registExcelFile(self, filename):
        if not os.path.exists(filename):
            return
        wb = openpyxl.load_workbook(filename)
        for sheetName in wb.sheetnames:
            ws = wb[sheetName]
            for row in range(1, ws.max_row+1):
                for col in range(1, ws.max_column+1):
                    cell = ws.cell(row=row, column=col)
                    value = str(cell.value).strip()
                    if value == '':
                        continue
                    hash = self.registCommon('excel', value, False)
                    self.dao.insertExcelFileTable(hash, filename, sheetName, row, col)

    def registWordFile(self, filename):
        #filename
        #page
        #line_no
        pass

    def registPdfFile(self, filename):
        #filename
        #page
        #line_no
        pass

    def registDirectory(self, dir):
        files = glob.glob(os.path.normpath(os.path.join(dir, '**/*')), recursive=True)
        for path in files:
            ext = os.path.splitext(path)[1].lower()
            if ext == '.txt':
                self.registTextFile(path)
            elif ext == '.xlsx' or ext == '.xlsm':
                # EXCEL
                self.registExcelFile(path)
            elif ext == '.docx':
                # WORD
                self.registWordFile(path)
            elif ext == '.pdf':
                # PDF
                self.registPdfFile(path)
            else:
                continue

    def searchCommon(self, keyword):
        return self.dao.selectTextTable(keyword)

    def searchWww(self, hash):
        return self.dao.selectWwwTable(hash)

    def searchSourceCode(self, hash):
        return self.dao.selectSourceFileTable(hash)

    def searchTextFile(self, hash):
        return self.dao.selectTextFileTable(hash)

    def searchExcelFile(self, hash):
        return self.dao.selectExcelFileTable(hash)

    def search(self, keyword):
        results = self.searchCommon(keyword)
        result_text = ''
        first = True
        for row in results:
            source_type = row[1]
            hash = row[2]
            if source_type == 'memo':
                if first:
                    first = False
                else:
                    result_text += '\n-------------------------\n'
                result_text += row[0]
            elif source_type == 'www':
                if first:
                    first = False
                else:
                    result_text += '\n-------------------------\n'
                html = self.searchWww(hash)
                result_text += html
            elif source_type == 'code':
                codes = self.searchSourceCode(hash)
                for code in codes:
                    if first:
                        first = False
                    else:
                        result_text += '\n-------------------------\n'
                    result_text += '{0} ({1}):\n{2}'.format(code[0], code[1], row[0])
            elif source_type == 'text':
                text = self.searchTextFile(hash)
                for line in text:
                    if first:
                        first = False
                    else:
                        result_text += '\n-------------------------\n'
                    result_text += '{0} {1}:\n{2}'.format(line[0], line[1], row[0])
            elif source_type == 'excel':
                cells = self.searchExcelFile(hash)
                for cell in cells:
                    if first:
                        first = False
                    else:
                        result_text += '\n-------------------------\n'
                    result_text += '{0} 「{1}」シート({2}):\n{3}'.format(cell[0], cell[1], '{0}{1}'.format(openpyxl.utils.get_column_letter(cell[2]), cell[3]), row[0])
            else:
                if first:
                    first = False
                else:
                    result_text += '\n-------------------------\n'
                result_text += row[0]

        return result_text

